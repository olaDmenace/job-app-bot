#!/usr/bin/env python3
"""
Data Export System for Job Application Bot
Exports job data from SQLite database to CSV and Excel formats
"""

import os
import csv
import sqlite3
from datetime import datetime
from typing import List, Dict, Optional

class JobDataExporter:
    """Handles exporting job data to various formats"""
    
    def __init__(self, db_path: str = "job_applications.db"):
        """
        Initialize the data exporter.
        
        Args:
            db_path (str): Path to the SQLite database
        """
        self.db_path = db_path
        
    def _get_database_connection(self):
        """Get database connection"""
        if not os.path.exists(self.db_path):
            raise FileNotFoundError(f"Database not found: {self.db_path}")
        return sqlite3.connect(self.db_path)
    
    def get_jobs_data(self, limit: Optional[int] = None, source: Optional[str] = None, 
                      days_back: Optional[int] = None) -> List[Dict]:
        """
        Retrieve jobs data from database.
        
        Args:
            limit (int, optional): Maximum number of jobs to retrieve
            source (str, optional): Filter by job source (e.g., 'adzuna', 'jsearch')
            days_back (int, optional): Only get jobs from last N days
            
        Returns:
            List[Dict]: List of job dictionaries
        """
        conn = self._get_database_connection()
        cursor = conn.cursor()
        
        # Build query with filters
        query = """
        SELECT job_source_id as id, title, company, location, salary, 
               date_posted as posted, tags, url, date_found, source, 
               description, is_remote
        FROM jobs 
        WHERE 1=1
        """
        params = []
        
        if source:
            query += " AND source = ?"
            params.append(source)
        
        if days_back:
            query += " AND date(date_found) >= date('now', '-{} days')".format(days_back)
        
        query += " ORDER BY date_found DESC"
        
        if limit:
            query += " LIMIT ?"
            params.append(limit)
        
        cursor.execute(query, params)
        columns = [description[0] for description in cursor.description]
        
        jobs = []
        for row in cursor.fetchall():
            job_dict = dict(zip(columns, row))
            jobs.append(job_dict)
        
        conn.close()
        return jobs
    
    def export_to_csv(self, filename: Optional[str] = None, **kwargs) -> str:
        """
        Export jobs data to CSV file.
        
        Args:
            filename (str, optional): Custom filename for export
            **kwargs: Additional filters (limit, source, days_back)
            
        Returns:
            str: Path to exported CSV file
        """
        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"jobs_export_{timestamp}.csv"
        
        # Ensure .csv extension
        if not filename.endswith('.csv'):
            filename += '.csv'
        
        # Get data from database
        jobs_data = self.get_jobs_data(**kwargs)
        
        if not jobs_data:
            print("No jobs found to export")
            return ""
        
        # Write to CSV
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = jobs_data[0].keys()
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(jobs_data)
        
        print(f"Exported {len(jobs_data)} jobs to {filename}")
        return filename
    
    def export_to_excel(self, filename: Optional[str] = None, **kwargs) -> str:
        """
        Export jobs data to Excel file.
        
        Args:
            filename (str, optional): Custom filename for export
            **kwargs: Additional filters (limit, source, days_back)
            
        Returns:
            str: Path to exported Excel file
        """
        try:
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("Excel export requires openpyxl. Install with: pip install openpyxl")
            return ""
        
        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"jobs_export_{timestamp}.xlsx"
        
        # Ensure .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        # Get data from database
        jobs_data = self.get_jobs_data(**kwargs)
        
        if not jobs_data:
            print("No jobs found to export")
            return ""
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Job Applications"
        
        # Add headers with styling
        headers = list(jobs_data[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for row_num, job in enumerate(jobs_data, 2):
            for col_num, header in enumerate(headers, 1):
                value = job.get(header, '')
                # Handle None values
                if value is None:
                    value = ''
                ws.cell(row=row_num, column=col_num, value=str(value))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(filename)
        print(f"Exported {len(jobs_data)} jobs to {filename}")
        return filename
    
    def get_export_summary(self) -> Dict:
        """
        Get summary of available data for export.
        
        Returns:
            Dict: Summary statistics
        """
        conn = self._get_database_connection()
        cursor = conn.cursor()
        
        # Total jobs
        cursor.execute("SELECT COUNT(*) FROM jobs")
        total_jobs = cursor.fetchone()[0]
        
        # Jobs by source
        cursor.execute("SELECT source, COUNT(*) FROM jobs GROUP BY source ORDER BY COUNT(*) DESC")
        jobs_by_source = dict(cursor.fetchall())
        
        # Recent jobs (last 7 days)
        cursor.execute("SELECT COUNT(*) FROM jobs WHERE date(date_found) >= date('now', '-7 days')")
        recent_jobs = cursor.fetchone()[0]
        
        # Date range
        cursor.execute("SELECT MIN(date(date_found)), MAX(date(date_found)) FROM jobs")
        date_range = cursor.fetchone()
        
        conn.close()
        
        return {
            'total_jobs': total_jobs,
            'jobs_by_source': jobs_by_source,
            'recent_jobs_7_days': recent_jobs,
            'date_range': {
                'earliest': date_range[0],
                'latest': date_range[1]
            }
        }
    
    def print_export_summary(self):
        """Print a formatted summary of available data"""
        summary = self.get_export_summary()
        
        print("\nJob Data Export Summary")
        print("=" * 40)
        print(f"Total jobs in database: {summary['total_jobs']}")
        print(f"Recent jobs (last 7 days): {summary['recent_jobs_7_days']}")
        
        if summary['date_range']['earliest']:
            print(f"Data range: {summary['date_range']['earliest']} to {summary['date_range']['latest']}")
        
        print("\nJobs by source:")
        for source, count in summary['jobs_by_source'].items():
            print(f"  {source}: {count} jobs")
        
        print("\nExport options:")
        print("  CSV: python -m data_exporter --csv")
        print("  Excel: python -m data_exporter --excel")
        print("  Recent only: python -m data_exporter --csv --days-back 7")
        print("  Specific source: python -m data_exporter --csv --source adzuna")

def main():
    """CLI interface for data exporter"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Export job data from database')
    parser.add_argument('--csv', action='store_true', help='Export to CSV format')
    parser.add_argument('--excel', action='store_true', help='Export to Excel format')
    parser.add_argument('--filename', type=str, help='Custom filename for export')
    parser.add_argument('--limit', type=int, help='Limit number of jobs to export')
    parser.add_argument('--source', type=str, help='Filter by job source (adzuna, jsearch, etc.)')
    parser.add_argument('--days-back', type=int, help='Only export jobs from last N days')
    parser.add_argument('--summary', action='store_true', help='Show data summary')
    
    args = parser.parse_args()
    
    exporter = JobDataExporter()
    
    if args.summary:
        exporter.print_export_summary()
        return
    
    if not args.csv and not args.excel:
        exporter.print_export_summary()
        return
    
    # Prepare export parameters
    export_params = {}
    if args.limit:
        export_params['limit'] = args.limit
    if args.source:
        export_params['source'] = args.source
    if args.days_back:
        export_params['days_back'] = args.days_back
    
    # Export to requested formats
    if args.csv:
        filename = exporter.export_to_csv(args.filename, **export_params)
        if filename:
            print(f"CSV export complete: {filename}")
    
    if args.excel:
        filename = exporter.export_to_excel(args.filename, **export_params)
        if filename:
            print(f"Excel export complete: {filename}")

if __name__ == "__main__":
    main()